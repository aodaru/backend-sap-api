"""
Servicio de Condiciones - VK12.

Implementa la lógica de negocio para la transacción SAP VK12
(modificación masiva de condiciones de precio).
"""

import io
import logging
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from fastapi import UploadFile

from models.responses import JobStatus, ValidationDetail

logger = logging.getLogger(__name__)

# --- Constantes de validación VK12 ---

REQUIRED_COLUMNS = [
    "MATERIAL",
    "UNIDAD_DE_MEDIDA",
    "IMPORTE",
    "GRUPO_ARTICULO",
    "ORG_VENTA",
    "CAN_DISTR",
    "SECTOR",
    "RAMO",
    "TIPO_MODIFICACION",
]

VALID_FLOWS = {
    "mat_orgvent_candistr",
    "orgvent_candistr_gpoart",
    "orgvent_candistr_sec_ramo_mat",
    "orgven_candist_sec_gpoart",
}

# Campos requeridos por cada flujo
FLOW_FIELDS: Dict[str, List[str]] = {
    "mat_orgvent_candistr": [
        "MATERIAL",
        "UNIDAD_DE_MEDIDA",
        "IMPORTE",
        "ORG_VENTA",
        "CAN_DISTR",
    ],
    "orgvent_candistr_gpoart": [
        "ORG_VENTA",
        "CAN_DISTR",
        "GRUPO_ARTICULO",
        "IMPORTE",
    ],
    "orgvent_candistr_sec_ramo_mat": [
        "ORG_VENTA",
        "CAN_DISTR",
        "SECTOR",
        "RAMO",
        "MATERIAL",
    ],
    "orgven_candist_sec_gpoart": [
        "ORG_VENTA",
        "CAN_DISTR",
        "SECTOR",
        "RAMO",
        "GRUPO_ARTICULO",
    ],
}

# Valores válidos por campo
VALID_ORG_VENTA = {"1000"}
VALID_CAN_DISTR = {"10", "20", "30", "40", "50"}
VALID_SECTOR = {"10", "20", "30", "40", "50"}
VALID_RAMO = {"ZDET", "ZCON", "ZPRO", "ZCOL", "ZMIN"}
VALID_UNIDAD_MEDIDA = {
    "UN", "KG", "ST", "L", "M", "M2", "M3", "LB", "GL",
    "OZ", "PC", "BX", "CA", "SET", "ROL", "PAR", "DOC",
    "TN", "TO", "G", "MG", "ML", "CL", "DL", "HL",
}


def _solo_numeros(texto: str) -> bool:
    """Verifica que el texto contenga solo caracteres numéricos."""
    texto = str(texto).strip()
    return texto.isdigit()


def _is_numeric(value: str) -> bool:
    """Verifica que el valor sea numérico (entero o decimal)."""
    try:
        float(value)
        return True
    except (ValueError, TypeError):
        return False


def _validate_field(
    field: str, value: Any, row_num: int
) -> Optional[str]:
    """
    Valida un campo individual del Excel VK12.

    Args:
        field: Nombre del campo.
        value: Valor del campo.
        row_num: Número de fila.

    Returns:
        Mensaje de error o None si es válido.
    """
    if value is None:
        return f"Fila {row_num}: '{field}' es obligatorio"

    value_str = str(value).strip()

    if not value_str or value_str.lower() == "nan":
        return f"Fila {row_num}: '{field}' es obligatorio"

    if field == "IMPORTE" and not _is_numeric(value_str):
        return f"Fila {row_num}: '{field}' debe ser numérico"

    if field == "MATERIAL" and not _solo_numeros(value_str):
        return (
            f"Fila {row_num}: '{field} - {value_str}' "
            "no debe contener letras ni caracteres especiales"
        )

    if field == "ORG_VENTA" and value_str not in VALID_ORG_VENTA:
        return f"Fila {row_num}: '{field}' debe ser '1000'"

    if field == "CAN_DISTR" and value_str not in VALID_CAN_DISTR:
        return (
            f"Fila {row_num}: '{field}' debe ser 10, 20, 30, 40 o 50"
        )

    if field == "SECTOR" and value_str not in VALID_SECTOR:
        return (
            f"Fila {row_num}: '{field} - {value_str}' "
            "debe ser 10, 20, 30, 40 o 50"
        )

    if field == "RAMO" and value_str.upper() not in VALID_RAMO:
        return (
            f"Fila {row_num}: '{field}' debe ser ZDET, ZCON, ZPRO, ZCOL o ZMIN"
        )

    if field == "GRUPO_ARTICULO":
        if len(value_str) != 9:
            return (
                f"Fila {row_num}: '{field} - {value_str}' "
                "debe tener 9 caracteres"
            )
        if not _solo_numeros(value_str):
            return (
                f"Fila {row_num}: '{field}' "
                "no debe contener letras ni caracteres especiales"
            )

    if (
        field == "UNIDAD_DE_MEDIDA"
        and value_str.upper() not in VALID_UNIDAD_MEDIDA
    ):
        return (
            f"Fila {row_num}: '{field}' No es una unidad de medida válida"
        )

    return None


def get_template_path() -> Path:
    """Retorna la ruta al template Excel de condiciones VK12."""
    return Path(__file__).parent.parent / "templates" / "condiciones_template.xlsx"


async def validate_excel(
    file: UploadFile,
) -> Tuple[bool, List[ValidationDetail], List[Dict[str, Any]]]:
    """
    Valida un archivo Excel de condiciones VK12.

    Args:
        file: Archivo Excel subido por el usuario.

    Returns:
        Tupla con (es_válido, lista_de_errores, datos_parseados).
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl no está instalado")

    errors: List[ValidationDetail] = []
    rows_data: List[Dict[str, Any]] = []

    try:
        content = await file.read()

        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active

        if ws is None:
            errors.append(
                ValidationDetail(
                    row=0, field="sheet", error="El archivo no tiene hojas"
                )
            )
            return False, errors, []

        # Leer headers (normalizar a mayúsculas y sin espacios)
        raw_headers = [
            cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))
        ]
        headers = [
            str(h).strip().upper().replace(" ", "_")
            for h in raw_headers
            if h is not None
        ]

        # Validar columnas requeridas
        missing_columns = [col for col in REQUIRED_COLUMNS if col not in headers]
        if missing_columns:
            errors.append(
                ValidationDetail(
                    row=1,
                    field="columns",
                    error=f"Columnas faltantes: {', '.join(missing_columns)}",
                )
            )
            return False, errors, []

        # Validar filas de datos
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            if all(cell is None for cell in row):
                continue  # Saltar filas vacías

            row_dict = dict(zip(headers, row))
            rows_data.append(row_dict)

            # Obtener flujo de esta fila
            flow_name = str(row_dict.get("TIPO_MODIFICACION", "")).strip().lower()

            # Validar que TIPO_MODIFICACION sea un flujo válido
            if flow_name not in VALID_FLOWS:
                errors.append(
                    ValidationDetail(
                        row=row_idx,
                        field="TIPO_MODIFICACION",
                        error=f"Flujo '{flow_name}' no válido",
                    )
                )
                continue

            # Validar campos requeridos según el flujo
            required_fields = FLOW_FIELDS.get(flow_name, [])
            for field in required_fields:
                error = _validate_field(field, row_dict.get(field), row_idx)
                if error:
                    errors.append(
                        ValidationDetail(
                            row=row_idx, field=field, error=error
                        )
                    )

        wb.close()

        is_valid = len(errors) == 0
        return is_valid, errors, rows_data

    except Exception as e:
        errors.append(
            ValidationDetail(
                row=0,
                field="file",
                error=f"Error al leer archivo: {e!s}",
            )
        )
        return False, errors, []


async def execute_vk12(
    file_data: List[Dict[str, Any]],
    job_id: str,
    credentials: Optional[Dict[str, str]] = None,
) -> Dict[str, Any]:
    """
    Ejecuta la transacción SAP VK12 con los datos proporcionados.

    En producción, esta función interactuaría con SAP GUI via win32com.
    En tests, se puede mockear esta función.

    La ejecución pasa por el sistema de cola de peticiones:
    1. Se encola la petición
    2. Se espera turno
    3. Se ejecuta con reintentos automáticos

    Args:
        file_data: Lista de diccionarios con los datos del Excel.
        job_id: ID del job en ejecución.
        credentials: Credenciales SAP (system, mandt, username, password, language).

    Returns:
        Diccionario con los resultados de la ejecución.
    """
    from services.costos_service import job_manager
    from services.queue_service import request_queue

    # Encolar la petición
    try:
        await request_queue.enqueue(
            job_id=job_id,
            transaction="VK12",
            user_id=credentials.get("username", "system") if credentials else "system",
        )
    except ValueError:
        # Cola llena - propagar error
        job_manager.update_job(
            job_id,
            JobStatus.FAILED,
            progress=0,
            results={"message": "Cola de peticiones llena"},
        )
        raise

    # Marcar como procesando
    job_manager.update_job(job_id, JobStatus.PROCESSING, progress=10)

    # Desencolar para procesar
    await request_queue.dequeue()

    # Función interna de ejecución
    async def _execute_sap() -> Dict[str, Any]:
        """Ejecuta la transacción SAP VK12."""
        # TODO: Implementar integración real con SAP GUI
        # Por ahora retorna mock de éxito
        return {
            "processed": len(file_data),
            "successful": len(file_data),
            "failed": 0,
            "message": "VK12 ejecutado exitosamente",
        }

    try:
        # Ejecutar con reintentos
        results = await request_queue.process_with_retries(
            job_id, _execute_sap
        )

        # Marcar como completada en la cola
        await request_queue.mark_completed(job_id, results)

        # Actualizar job manager
        job_manager.update_job(
            job_id,
            JobStatus.COMPLETED,
            progress=100,
            results={
                "processed": len(file_data),
                "successful": len(file_data),
                "failed": 0,
                "message": results.get("message", "VK12 ejecutado exitosamente"),
            },
        )

        return {
            "processed": len(file_data),
            "successful": len(file_data),
            "failed": 0,
        }

    except Exception as e:
        # Marcar como fallida en la cola
        error_msg = str(e)
        await request_queue.mark_failed(job_id, error_msg)

        # Actualizar job manager
        job_manager.update_job(
            job_id,
            JobStatus.FAILED,
            progress=0,
            results={"message": error_msg},
        )

        raise

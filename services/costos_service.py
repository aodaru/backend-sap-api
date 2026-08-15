"""
Servicio de Costos - ME12.

Implementa la lógica de negocio para la transacción SAP ME12
(modificación de precios de material).
"""

import uuid
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, cast

from fastapi import UploadFile

from models.responses import JobStatus, ValidationDetail

# Columnas requeridas en el template Excel
REQUIRED_COLUMNS = [
    "Material",
    "Proveedor",
    "Org_Compras",
    "Tipo_Info",
    "Tipo_Condicion",
    "Nuevo_Precio",
    "Moneda",
    "Unidad_Precio",
    "Unidad_Medida",
    "Valido_Desde",
    "Valido_Hasta",
]


class JobManager:
    """Gestiona jobs en memoria para ejecución de transacciones SAP."""

    def __init__(self) -> None:
        self._jobs: Dict[str, Dict[str, Any]] = {}

    def create_job(self) -> str:
        """Crea un nuevo job y retorna su ID."""
        job_id = str(uuid.uuid4())
        self._jobs[job_id] = {
            "job_id": job_id,
            "status": JobStatus.PENDING,
            "progress": 0,
            "results": None,
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        return job_id

    def update_job(
        self,
        job_id: str,
        status: JobStatus,
        progress: int = 0,
        results: Optional[Dict[str, Any]] = None,
    ) -> None:
        """Actualiza el estado de un job."""
        if job_id not in self._jobs:
            raise KeyError(f"Job {job_id} not found")
        self._jobs[job_id]["status"] = status
        self._jobs[job_id]["progress"] = progress
        if results is not None:
            self._jobs[job_id]["results"] = results

    def get_job(self, job_id: str) -> Optional[Dict[str, Any]]:
        """Retorna el estado de un job o None si no existe."""
        return self._jobs.get(job_id)


# Instancia global del JobManager
job_manager = JobManager()


def get_template_path() -> Path:
    """Retorna la ruta al template Excel de costos."""
    return Path(__file__).parent.parent / "templates" / "costos_template.xlsx"


def _apply_me12_date_defaults(row_data: Dict[str, Any]) -> None:
    """Completa fechas opcionales de ME12 en el formato contractual.

    Excel puede entregar fechas como objetos ``date``/``datetime`` y las
    celdas vacías pueden llegar como ``None`` o como texto en blanco. Solo se
    modifican las dos columnas de vigencia; los demás valores permanecen
    intactos.
    """
    date_fields = {
        "Valido_Desde": date.today().strftime("%d.%m.%Y"),
        "Valido_Hasta": "31.12.9999",
    }

    for field, default in date_fields.items():
        value = row_data.get(field)
        if value is None or (isinstance(value, str) and not value.strip()):
            row_data[field] = default
        elif isinstance(value, (datetime, date)):
            row_data[field] = value.strftime("%d.%m.%Y")


async def validate_excel(
    file: UploadFile,
) -> Tuple[bool, List[ValidationDetail], List[Dict[str, Any]]]:
    """
    Valida un archivo Excel de costos.

    Args:
        file: Archivo Excel subido por el usuario.

    Returns:
        Tupla con (es_válido, lista_de_errores, datos_parseados).
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl no está instalado")

    errors: List[ValidationDetail] = []
    rows_data: List[Dict[str, Any]] = []

    try:
        content = await file.read()
        import io

        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active

        if ws is None:
            errors.append(
                ValidationDetail(row=0, field="sheet", error="El archivo no tiene hojas")
            )
            return False, errors, []

        # Leer headers
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        headers = [h for h in headers if h is not None]

        # Validar columnas requeridas
        missing_columns = [col for col in REQUIRED_COLUMNS if col not in headers]
        if missing_columns:
            errors.append(
                ValidationDetail(
                    row=1,
                    field="columns",
                    error=f"Columnas faltantes: {', '.join(missing_columns)}",
                )
            )
            return False, errors, []

        # Validar filas de datos
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(cell is None for cell in row):
                continue  # Saltar filas vacías

            row_dict = cast(Dict[str, Any], dict(zip(headers, row)))
            _apply_me12_date_defaults(row_dict)
            rows_data.append(row_dict)

            # Validar que no haya campos vacíos
            for col in REQUIRED_COLUMNS:
                if col in row_dict and row_dict[col] is None:
                    errors.append(
                        ValidationDetail(
                            row=row_idx, field=col, error="Campo requerido vacío"
                        )
                    )

            # Validar tipos numéricos
            if "Nuevo_Precio" in row_dict and row_dict["Nuevo_Precio"] is not None:
                try:
                    float(row_dict["Nuevo_Precio"])
                except (ValueError, TypeError):
                    errors.append(
                        ValidationDetail(
                            row=row_idx,
                            field="Nuevo_Precio",
                            error="Debe ser un número",
                        )
                    )

        wb.close()

        is_valid = len(errors) == 0
        return is_valid, errors, rows_data

    except Exception as e:
        errors.append(
            ValidationDetail(row=0, field="file", error="No se pudo leer el archivo")
        )
        return False, errors, []


async def execute_me12(
    file_data: List[Dict[str, Any]], job_id: str
) -> Dict[str, Any]:
    """
    Ejecuta la transacción SAP ME12 con los datos proporcionados.

    En producción, esta función interactuaría con SAP GUI via win32com.
    En tests, se puede mockear esta función.

    La ejecución pasa por el sistema de cola de peticiones:
    1. Se encola la petición
    2. Se espera turno
    3. Se ejecuta con reintentos automáticos

    Args:
        file_data: Lista de diccionarios con los datos del Excel.
        job_id: ID del job en ejecución.

    Returns:
        Diccionario con los resultados de la ejecución.
    """
    from services.queue_service import request_queue

    try:
        job_manager.update_job(job_id, JobStatus.QUEUED, progress=0)
        async def _execute_sap() -> Dict[str, Any]:
            job_manager.update_job(job_id, JobStatus.PROCESSING, progress=10)
            from services.sap_executor import sap_executor
            return (await sap_executor.execute("ME12", file_data)).as_dict()
        results = await request_queue.run_job(job_id, "ME12", _execute_sap)
    except ValueError:
        # Cola llena - propagar error
        job_manager.update_job(
            job_id,
            JobStatus.FAILED,
            progress=0,
            results={"message": "Cola de peticiones llena"},
        )
        raise
    except Exception as e:
        from services.sap_errors import safe_exception
        safe_error = safe_exception(e)
        error_msg = getattr(safe_error, "public_message", "No se pudo completar ME12")
        job_manager.update_job(
            job_id,
            JobStatus.TIMEOUT if getattr(safe_error, "code", "") in {"execution_timeout", "queue_wait_timeout"} else JobStatus.FAILED,
            progress=0,
            results={"message": error_msg},
        )
        raise safe_error

    try:
        # Actualizar job manager
        job_manager.update_job(
            job_id,
            JobStatus.COMPLETED,
            progress=100,
            results={
                "processed": len(file_data),
                "successful": results["successful"],
                "failed": results["failed"],
                "message": results.get("message", "ME12 ejecutado exitosamente"),
                "rows": results.get("rows", []),
            },
        )

        return {
            "processed": len(file_data),
            "successful": results["successful"],
            "failed": results["failed"],
            "rows": results.get("rows", []),
        }

    except Exception as e:
        error_msg = getattr(e, "public_message", "No se pudo completar ME12")

        # Actualizar job manager
        job_manager.update_job(
            job_id,
            JobStatus.TIMEOUT if getattr(e, "code", "") in {"execution_timeout", "queue_wait_timeout"} else JobStatus.FAILED,
            progress=0,
            results={"message": error_msg},
        )

        raise

"""
Configuración de tests - Fixtures compartidos.

Proporciona fixtures reutilizables para toda la suite de tests,
incluyendo el cliente de prueba FastAPI.
"""

import io
from typing import Generator

import pytest
from fastapi.testclient import TestClient

from main import app


class _MockControl:
    """Mock de un control SAP GUI COM — acepta asignaciones de .Text, .SetFocus, etc."""

    def __init__(self):
        self.Text = ""
        self.CaretPosition = 0

    def SetFocus(self):
        pass

    def SendVKey(self, code):
        pass

    def resizeWorkingPane(self, w, h, flag):
        pass


class _TestSapSession:
    """Sesión falsa: permite probar navegación sin instalar SAP GUI."""

    def findById(self, control_id):
        if control_id == "wnd[1]":
            raise Exception("no popup")
        return _MockControl()

    def execute_transaction(self, transaction, row):
        return None


class _TestSapProvider:
    def acquire(self, credentials=None):
        return _TestSapSession()

    def release(self, session):
        return None


@pytest.fixture(autouse=True)
def fake_sap_provider(monkeypatch, request):
    """Aísla la suite de SAP real sin convertir el proveedor nulo en éxito."""
    if request.node.get_closest_marker("sap_integration_disabled"):
        return
    from services.sap_executor import SapTransactionExecutor

    monkeypatch.setattr("services.sap_executor.get_settings", lambda: type("TestSettings", (), {
        "sap_integration_enabled": True,
        "max_retries": 0,
        "sap_execution_timeout": 120,
        "sap_retry_backoff": 0,
    })())
    monkeypatch.setattr("services.sap_executor.sap_executor", SapTransactionExecutor(_TestSapProvider()))


@pytest.fixture(scope="module")
def client() -> TestClient:
    """
    Fixture que retorna un cliente de prueba FastAPI.

    Scope: module - se crea una vez por módulo de test para mejor rendimiento.

    Returns:
        TestClient: Cliente HTTP para hacer requests contra la app.
    """
    return TestClient(app)


@pytest.fixture(scope="module")
def valid_api_key() -> str:
    """
    Fixture que retorna una API key válida para tests.

    Returns:
        str: API key de prueba.
    """
    return "mi-api-key-secreta"


@pytest.fixture
def valid_excel_file() -> Generator[io.BytesIO, None, None]:
    """
    Fixture que retorna un archivo Excel válido para tests.

    Returns:
        Generator con buffer BytesIO conteniendo un Excel válido.
    """
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Costos ME12"

    headers = [
        "Material",
        "Proveedor",
        "Org_Compras",
        "Tipo_Info",
        "Tipo_Condicion",
        "Nuevo_Precio",
        "Moneda",
        "Unidad_Precio",
        "Unidad_Medida",
        "Valido_Desde",
        "Valido_Hasta",
    ]
    ws.append(headers)

    # Fila de datos válidos
    ws.append(
        [
            "MAT001",
            "PROV001",
            "1000",
            "0",
            "PB00",
            100.50,
            "MXN",
            "ST",
            "ST",
            "20260101",
            "20261231",
        ]
    )

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    buffer.name = "test_valid.xlsx"

    yield buffer


@pytest.fixture
def invalid_excel_file_missing_columns() -> Generator[io.BytesIO, None, None]:
    """
    Fixture que retorna un Excel inválido (columnas faltantes).

    Returns:
        Generator con buffer BytesIO conteniendo un Excel inválido.
    """
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Costos ME12"

    # Solo algunas columnas (faltan varias)
    ws.append(["Material", "Proveedor"])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    buffer.name = "test_invalid.xlsx"

    yield buffer


@pytest.fixture
def invalid_excel_file_bad_types() -> Generator[io.BytesIO, None, None]:
    """
    Fixture que retorna un Excel con tipos incorrectos.

    Returns:
        Generator con buffer BytesIO conteniendo un Excel con errores de tipo.
    """
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Costos ME12"

    headers = [
        "Material",
        "Proveedor",
        "Org_Compras",
        "Tipo_Info",
        "Tipo_Condicion",
        "Nuevo_Precio",
        "Moneda",
        "Unidad_Precio",
        "Unidad_Medida",
        "Valido_Desde",
        "Valido_Hasta",
    ]
    ws.append(headers)

    # Nuevo_Precio con valor no numérico
    ws.append(
        [
            "MAT001",
            "PROV001",
            "1000",
            "0",
            "PB00",
            "no_es_numero",  # Error de tipo
            "MXN",
            "ST",
            "ST",
            "20260101",
            "20261231",
        ]
    )

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    buffer.name = "test_bad_types.xlsx"

    yield buffer


# --- Fixtures para Condiciones VK12 ---


@pytest.fixture
def valid_condiciones_excel() -> Generator[io.BytesIO, None, None]:
    """
    Fixture que retorna un archivo Excel válido para condiciones VK12.

    Incluye datos válidos para el flujo mat_orgvent_candistr.

    Returns:
        Generator con buffer BytesIO conteniendo un Excel válido.
    """
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "VK12 Condiciones"

    headers = [
        "MATERIAL", "UNIDAD_DE_MEDIDA", "IMPORTE", "GRUPO_ARTICULO",
        "ORG_VENTA", "CAN_DISTR", "SECTOR", "RAMO", "TIPO_MODIFICACION",
    ]
    ws.append(headers)

    # Fila válida - flujo mat_orgvent_candistr
    ws.append([
        "12345678", "UN", 100.50, "123456789",
        "1000", "10", "10", "ZDET", "mat_orgvent_candistr",
    ])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    buffer.name = "test_condiciones_valid.xlsx"

    yield buffer


@pytest.fixture
def invalid_condiciones_excel_missing_columns() -> Generator[io.BytesIO, None, None]:
    """
    Fixture que retorna un Excel inválido (columnas faltantes) para condiciones.

    Returns:
        Generator con buffer BytesIO conteniendo un Excel inválido.
    """
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "VK12 Condiciones"

    # Solo algunas columnas (faltan varias)
    ws.append(["MATERIAL", "IMPORTE"])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    buffer.name = "test_condiciones_invalid.xlsx"

    yield buffer


@pytest.fixture
def invalid_condiciones_excel_bad_flow() -> Generator[io.BytesIO, None, None]:
    """
    Fixture que retorna un Excel con flujo inválido para condiciones.

    Returns:
        Generator con buffer BytesIO conteniendo un Excel con flujo inválido.
    """
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "VK12 Condiciones"

    headers = [
        "MATERIAL", "UNIDAD_DE_MEDIDA", "IMPORTE", "GRUPO_ARTICULO",
        "ORG_VENTA", "CAN_DISTR", "SECTOR", "RAMO", "TIPO_MODIFICACION",
    ]
    ws.append(headers)

    # Fila con flujo inválido
    ws.append([
        "12345678", "UN", 100.50, "123456789",
        "1000", "10", "10", "ZDET", "flujo_inexistente",
    ])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    buffer.name = "test_condiciones_bad_flow.xlsx"

    yield buffer


@pytest.fixture
def invalid_condiciones_excel_bad_types() -> Generator[io.BytesIO, None, None]:
    """
    Fixture que retorna un Excel con tipos incorrectos para condiciones.

    Returns:
        Generator con buffer BytesIO conteniendo un Excel con errores de tipo.
    """
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "VK12 Condiciones"

    headers = [
        "MATERIAL", "UNIDAD_DE_MEDIDA", "IMPORTE", "GRUPO_ARTICULO",
        "ORG_VENTA", "CAN_DISTR", "SECTOR", "RAMO", "TIPO_MODIFICACION",
    ]
    ws.append(headers)

    # Fila con IMPORTE no numérico
    ws.append([
        "12345678", "UN", "no_es_numero", "123456789",
        "1000", "10", "10", "ZDET", "mat_orgvent_candistr",
    ])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    buffer.name = "test_condiciones_bad_types.xlsx"

    yield buffer
